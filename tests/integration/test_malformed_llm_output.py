from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from buratino.app import VerificationApp
from buratino.audit.service import AuditService
from buratino.llm.prompt_loader import PromptLoader
from buratino.models.domain import DocumentSummary, EventRecord, PhrRecord
from buratino.target_builder.service import TargetBuilder
from buratino.verifier.document_ranking import DocumentRankingService
from buratino.verifier.event_verifier import EventVerifier
from buratino.verifier.phr_verifier import PhrVerifier
from conftest import create_prompt_assets


class FakeEventRepository:
    def get_event(self, event_id: int) -> EventRecord:
        return EventRecord(
            event_id=event_id,
            event_name="Событие",
            event_description="Описание",
            planned_value=0,
            planned_unit="шт",
        )

    def get_event_phr(self, event_id: int) -> PhrRecord:
        return PhrRecord(event_id=event_id, phr_name="ПХР", phr_value_2025=0, phr_unit="шт")


class FakeSummaryRepository:
    def list_event_documents(self, event_id: int) -> list[DocumentSummary]:
        return [
            DocumentSummary(
                document_id="1",
                file_name="doc.pdf",
                evidence_text="ocr text",
                evidence_source="ocr",
                ocr_text="ocr text",
                summary_text="summary text",
                ocr_parts=("ocr text",),
            )
        ]


class SequencedLlmClient:
    def __init__(self, responses: list[str]) -> None:
        self._responses = responses
        self.prompts: list[str] = []

    def generate_json(self, *, model: str, prompt: str) -> str:
        self.prompts.append(prompt)
        return self._responses.pop(0)


def test_malformed_json_retry_can_recover(tmp_path: Path) -> None:
    app, llm = _build_app(
        tmp_path,
        [
            "{bad json",
            _event_result(True),
            _audit_result("подтверждено", "подтверждено", ["doc.pdf"]),
        ],
    )

    artifacts = app.verify(
        event_id=1,
        output_dir=tmp_path / "output",
        primary_model="primary",
        audit_model="audit",
        export_xlsx=False,
    )

    assert artifacts.report.event_fact_status == "подтверждено"
    assert artifacts.report.error_stage is None
    assert len(llm.prompts) == 2
    assert "raw_response" in llm.prompts[1]


def test_final_malformed_json_is_saved_in_report_and_xlsx(tmp_path: Path) -> None:
    app, _ = _build_app(
        tmp_path,
        [
            "{bad json",
            "{still bad",
            "",
            _audit_result("не подтверждено", "подтверждено", []),
        ],
    )

    artifacts = app.verify(
        event_id=1,
        output_dir=tmp_path / "output",
        primary_model="primary",
        audit_model="audit",
        export_xlsx=True,
    )

    assert artifacts.report.event_fact_status == "не подтверждено"
    assert artifacts.report.event_documents[0].error_stage == "doc_level"
    assert artifacts.report.error_stage == "doc_level"
    assert artifacts.report.error_type == "empty_response"
    assert artifacts.report.prompt_name == "event_fact_summary.md"
    assert artifacts.xlsx_path is not None

    workbook = load_workbook(artifacts.xlsx_path, read_only=True)
    summary_sheet = workbook["summary"]
    summary_rows = {row[0]: row[1] for row in summary_sheet.iter_rows(min_row=2, values_only=True)}
    assert summary_rows["error_stage"] == "doc_level"
    assert summary_rows["error_type"] == "empty_response"

    docs_sheet = workbook["documents"]
    headers = next(docs_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    first_doc = next(docs_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    header_index = {name: idx for idx, name in enumerate(headers)}
    assert "error_stage" in header_index
    assert "raw_response_preview" in header_index
    assert first_doc[header_index["error_stage"]] == "doc_level"


def test_batch_xlsx_includes_new_diagnostic_columns(tmp_path: Path) -> None:
    app, _ = _build_app(
        tmp_path,
        [
            _event_result(False),
            _audit_result("не подтверждено", "подтверждено", []),
        ],
    )

    artifacts = app.verify(
        event_id=1,
        output_dir=tmp_path / "output",
        primary_model="primary",
        audit_model="audit",
        export_xlsx=False,
    )

    from buratino.report.batch_xlsx_exporter import BatchResult, BatchXlsxExporter

    batch_path = BatchXlsxExporter(tmp_path / "batch.xlsx").export(
        [BatchResult(input_event_id=1, status="ok", report=artifacts.report, json_path=artifacts.json_path)]
    )

    workbook = load_workbook(batch_path, read_only=True)
    headers = next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))
    assert "event_diagnostic_reasoning" in headers
    assert "error_stage" in headers
    assert "ranking_selected_file_names" in headers


def _build_app(tmp_path: Path, responses: list[str]) -> tuple[VerificationApp, SequencedLlmClient]:
    prompts_dir = tmp_path / "prompts"
    prompts_dir.mkdir()
    create_prompt_assets(prompts_dir)
    llm = SequencedLlmClient(responses)
    prompt_loader = PromptLoader(prompts_dir)
    app = VerificationApp(
        event_repository=FakeEventRepository(),
        summary_repository=FakeSummaryRepository(),
        target_builder=TargetBuilder(prompt_loader, llm, "primary"),
        document_ranking_service=DocumentRankingService(prompt_loader, llm, "ranking"),
        event_verifier=EventVerifier(prompt_loader, llm, "primary"),
        phr_verifier=PhrVerifier(prompt_loader, llm, "primary"),
        audit_service=AuditService(prompt_loader, llm, "audit"),
    )
    return app, llm


def _event_result(confirmed: bool) -> str:
    return json.dumps(
        {
            "document_id": "1",
            "file_name": "doc.pdf",
            "fact_status": "подтверждено" if confirmed else "не подтверждено",
            "reasoning": "Подтверждение найдено." if confirmed else "Подтверждение не найдено.",
            "matched_action": "Событие" if confirmed else None,
            "matched_subject": "Описание" if confirmed else None,
            "completion_signal": "выполнено" if confirmed else None,
            "observed_value": None,
            "observed_unit": None,
            "comparison_result": "not_applicable" if confirmed else "insufficient_data",
            "evidence_quote": "выполнено" if confirmed else None,
            "reasoning_trace": {
                "reason_codes": ["mentions_completion_fact"] if confirmed else ["insufficient_evidence"],
                "evidence_items": [
                    {
                        "quote": "выполнено",
                        "page": None,
            "source": "ocr",
                        "why_relevant": "Direct evidence.",
                    }
                ]
                if confirmed
                else [],
                "missing_requirements": [] if confirmed else ["explicit evidence"],
                "short_rationale": "trace",
                "confidence": "high" if confirmed else "low",
            },
        }
    )


def _audit_result(event_status: str, phr_status: str, supporting_files: list[str]) -> str:
    return json.dumps(
        {
            "audit_result": "pass",
            "rule_violations": [],
            "final_event_fact_status": event_status,
            "final_phr_fact_status": phr_status,
            "final_supporting_files": supporting_files,
        }
    )
