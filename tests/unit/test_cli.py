from __future__ import annotations

import threading
import time
from dataclasses import dataclass
from pathlib import Path

from buratino.cli.main import build_parser, parse_verify_command, read_event_ids, run, validate_event_id
from buratino.models.errors import BuratinoError
from buratino.models.contracts import VerificationReport
from conftest import create_prompt_assets


def clear_env(monkeypatch) -> None:
    for key in (
        "RANKING_MODEL",
        "PRIMARY_MODEL",
        "AUDIT_MODEL",
        "DATABASE_URL",
        "MAIN_DATABASE_URL",
        "RUNTIME_DATABASE_URL",
        "MAIN_DB_SCHEMA",
        "RUNTIME_DB_SCHEMA",
        "PROMPTS_DIR",
        "OUTPUT_DIR",
        "LLM_BACKEND",
        "LLM_API_BASE",
        "LLM_API_KEY",
        "LLM_TIMEOUT_SECONDS",
        "LLM_TEMPERATURE",
        "LLM_MAX_TOKENS",
        "EVENT_MAX_CONCURRENCY",
        "RANKING_BATCH_SIZE",
        "RANKING_SUMMARY_MAX_CHARS",
        "MAX_DOCUMENTS_TO_ANALYZE",
        "OCR_CHUNK_MAX_CHARS",
        "OCR_CHUNK_OVERLAP_CHARS",
        "OCR_CHUNK_MAX_CHUNKS",
        "EVIDENCE_SOURCE_MODE",
        "CONFIRMING_RELATION_MAX_TEXT_CHARS",
        "CONFIRMING_RELATION_BATCH_SIZE",
        "EVIDENCE_TRACE_ENABLED",
        "REASONING_TRACE_MODE",
        "REASONING_TRACE_MAX_ITEMS",
        "SHORT_RATIONALE_MAX_CHARS",
        "EVIDENCE_QUOTE_MAX_CHARS",
    ):
        monkeypatch.delenv(key, raising=False)


@dataclass
class FakeArtifacts:
    report: VerificationReport
    json_path: Path
    xlsx_path: Path | None


class FakeApp:
    def verify(
        self,
        *,
        event_id: int,
        output_dir: Path,
        primary_model: str,
        audit_model: str,
        export_xlsx: bool,
        max_documents_to_analyze: int | None = None,
    ) -> FakeArtifacts:
        json_path = output_dir / f"event_{event_id}.json"
        json_path.write_text("{}", encoding="utf-8")
        xlsx_path = output_dir / f"event_{event_id}.xlsx" if export_xlsx else None
        return FakeArtifacts(
            report=VerificationReport(
                event_id=event_id,
                event_name="test",
                event_type="qualitative",
                event_fact_status="не подтверждено",
                phr_fact_status="не подтверждено",
                event_primary_file=None,
                phr_primary_file=None,
                logic_is_valid=True,
                primary_model=primary_model,
                audit_model=audit_model,
                event_reasoning="none",
                phr_reasoning="none",
            ),
            json_path=json_path,
            xlsx_path=xlsx_path,
        )


class ControlledFakeApp:
    def __init__(
        self,
        *,
        active_counter: dict[str, int],
        max_seen_counter: dict[str, int],
        lock: threading.Lock,
        failures: set[int] | None = None,
        delays: dict[int, float] | None = None,
    ) -> None:
        self._active_counter = active_counter
        self._max_seen_counter = max_seen_counter
        self._lock = lock
        self._failures = failures or set()
        self._delays = delays or {}

    def verify(
        self,
        *,
        event_id: int,
        output_dir: Path,
        primary_model: str,
        audit_model: str,
        export_xlsx: bool,
        max_documents_to_analyze: int | None = None,
    ) -> FakeArtifacts:
        with self._lock:
            self._active_counter["value"] += 1
            self._max_seen_counter["value"] = max(
                self._max_seen_counter["value"],
                self._active_counter["value"],
            )
        try:
            time.sleep(self._delays.get(event_id, 0.02))
            if event_id in self._failures:
                raise BuratinoError(f"boom-{event_id}")
            json_path = output_dir / f"event_{event_id}.json"
            json_path.write_text("{}", encoding="utf-8")
            return FakeArtifacts(
                report=VerificationReport(
                    event_id=event_id,
                    event_name=f"event-{event_id}",
                    event_type="qualitative",
                    event_fact_status="не подтверждено",
                    phr_fact_status="не подтверждено",
                    event_primary_file=None,
                    phr_primary_file=None,
                    logic_is_valid=True,
                    primary_model=primary_model,
                    audit_model=audit_model,
                    event_reasoning="none",
                    phr_reasoning="none",
                ),
                json_path=json_path,
                xlsx_path=None,
            )
        finally:
            with self._lock:
                self._active_counter["value"] -= 1


def test_validate_event_id_requires_integer() -> None:
    try:
        validate_event_id("abc")
    except ValueError as exc:
        assert str(exc) == "event_id must be an integer."
    else:
        raise AssertionError("Expected ValueError for non-integer event_id")


def test_validate_event_id_requires_positive_integer() -> None:
    try:
        validate_event_id("0")
    except ValueError as exc:
        assert str(exc) == "event_id must be a positive integer."
    else:
        raise AssertionError("Expected ValueError for non-positive event_id")


def test_run_verify_returns_error_when_config_is_missing(
    capsys,
    monkeypatch,
    tmp_path: Path,
) -> None:
    clear_env(monkeypatch)
    monkeypatch.chdir(tmp_path)
    exit_code = run(["verify", "123"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "PRIMARY_MODEL" in captured.err


def test_run_verify_accepts_valid_input(
    monkeypatch,
    tmp_path: Path,
    capsys,
) -> None:
    clear_env(monkeypatch)
    prompts_dir = tmp_path / "prompts"
    prompts_dir.mkdir()
    create_prompt_assets(prompts_dir)

    monkeypatch.setenv("PRIMARY_MODEL", "primary")
    monkeypatch.setenv("AUDIT_MODEL", "audit")
    monkeypatch.setenv("RANKING_MODEL", "ranking")
    monkeypatch.setenv("MAIN_DATABASE_URL", "postgresql://main")
    monkeypatch.setenv("RUNTIME_DATABASE_URL", "postgresql://runtime")
    monkeypatch.setenv("PROMPTS_DIR", str(prompts_dir))
    monkeypatch.setenv("OUTPUT_DIR", str(tmp_path / "output"))
    monkeypatch.setattr("buratino.cli.main.build_app", lambda settings: FakeApp())

    exit_code = run(["verify", "123", "--xlsx"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "event_123.json" in captured.out
    assert "event_123.xlsx" in captured.out
    assert "Starting verification: event_id=123 primary_model=primary" in captured.err


def test_run_verify_writes_and_overwrites_log_file(
    monkeypatch,
    tmp_path: Path,
    capsys,
) -> None:
    clear_env(monkeypatch)
    prompts_dir = tmp_path / "prompts"
    prompts_dir.mkdir()
    create_prompt_assets(prompts_dir)
    output_dir = tmp_path / "output"

    monkeypatch.setenv("PRIMARY_MODEL", "primary")
    monkeypatch.setenv("AUDIT_MODEL", "audit")
    monkeypatch.setenv("RANKING_MODEL", "ranking")
    monkeypatch.setenv("MAIN_DATABASE_URL", "postgresql://main")
    monkeypatch.setenv("RUNTIME_DATABASE_URL", "postgresql://runtime")
    monkeypatch.setenv("PROMPTS_DIR", str(prompts_dir))
    monkeypatch.setenv("OUTPUT_DIR", str(output_dir))
    monkeypatch.setattr("buratino.cli.main.build_app", lambda settings: FakeApp())

    exit_code = run(["verify", "123"])
    capsys.readouterr()
    assert exit_code == 0

    log_path = output_dir / "buratino.log"
    first_content = log_path.read_text(encoding="utf-8")
    assert "Starting verification: event_id=123 primary_model=primary" in first_content

    log_path.write_text("stale content", encoding="utf-8")

    exit_code = run(["verify", "123"])
    capsys.readouterr()
    assert exit_code == 0

    second_content = log_path.read_text(encoding="utf-8")
    assert "stale content" not in second_content
    assert "Starting verification: event_id=123 primary_model=primary" in second_content


def test_parse_verify_command_uses_validated_event_id(tmp_path: Path) -> None:
    parser = build_parser()
    args = parser.parse_args(["verify", "77", "--output-dir", str(tmp_path), "--xlsx"])

    command = parse_verify_command(args)

    assert command.event_id == 77
    assert command.output_dir == tmp_path
    assert command.export_xlsx is True


def test_read_event_ids_ignores_empty_lines_and_comments(tmp_path: Path) -> None:
    ids_file = tmp_path / "ids.txt"
    ids_file.write_text("123\n\n# comment\n456\n", encoding="utf-8")

    assert read_event_ids(ids_file) == [123, 456]


def test_run_verify_list_writes_batch_xlsx(
    monkeypatch,
    tmp_path: Path,
    capsys,
) -> None:
    clear_env(monkeypatch)
    prompts_dir = tmp_path / "prompts"
    prompts_dir.mkdir()
    create_prompt_assets(prompts_dir)
    ids_file = tmp_path / "ids.txt"
    ids_file.write_text("123\n456\n", encoding="utf-8")
    xlsx_path = tmp_path / "batch.xlsx"

    monkeypatch.setenv("PRIMARY_MODEL", "primary")
    monkeypatch.setenv("AUDIT_MODEL", "audit")
    monkeypatch.setenv("RANKING_MODEL", "ranking")
    monkeypatch.setenv("MAIN_DATABASE_URL", "postgresql://main")
    monkeypatch.setenv("RUNTIME_DATABASE_URL", "postgresql://runtime")
    monkeypatch.setenv("PROMPTS_DIR", str(prompts_dir))
    monkeypatch.setenv("OUTPUT_DIR", str(tmp_path / "output"))
    monkeypatch.setattr("buratino.cli.main.build_app", lambda settings: FakeApp())

    exit_code = run(["verify-list", str(ids_file), "--xlsx", str(xlsx_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert xlsx_path.exists()
    assert "Processed ids: 2" in captured.out
    assert "Successful: 2" in captured.out
    assert (tmp_path / "output" / "buratino.log").exists()


def test_run_verify_list_uses_event_max_concurrency_and_keeps_input_order(
    monkeypatch,
    tmp_path: Path,
    capsys,
) -> None:
    clear_env(monkeypatch)
    prompts_dir = tmp_path / "prompts"
    prompts_dir.mkdir()
    create_prompt_assets(prompts_dir)
    ids_file = tmp_path / "ids.txt"
    ids_file.write_text("101\n102\n103\n104\n", encoding="utf-8")
    xlsx_path = tmp_path / "batch.xlsx"

    monkeypatch.setenv("PRIMARY_MODEL", "primary")
    monkeypatch.setenv("AUDIT_MODEL", "audit")
    monkeypatch.setenv("RANKING_MODEL", "ranking")
    monkeypatch.setenv("MAIN_DATABASE_URL", "postgresql://main")
    monkeypatch.setenv("RUNTIME_DATABASE_URL", "postgresql://runtime")
    monkeypatch.setenv("PROMPTS_DIR", str(prompts_dir))
    monkeypatch.setenv("OUTPUT_DIR", str(tmp_path / "output"))
    monkeypatch.setenv("EVENT_MAX_CONCURRENCY", "3")

    active_counter = {"value": 0}
    max_seen_counter = {"value": 0}
    lock = threading.Lock()
    monkeypatch.setattr(
        "buratino.cli.main.build_app",
        lambda settings: ControlledFakeApp(
            active_counter=active_counter,
            max_seen_counter=max_seen_counter,
            lock=lock,
            delays={101: 0.05, 102: 0.01, 103: 0.03, 104: 0.02},
        ),
    )

    exit_code = run(["verify-list", str(ids_file), "--xlsx", str(xlsx_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert xlsx_path.exists()
    assert "Processed ids: 4" in captured.out
    assert max_seen_counter["value"] <= 3

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == [101, 102, 103, 104]


def test_run_verify_list_continues_after_item_error(
    monkeypatch,
    tmp_path: Path,
    capsys,
) -> None:
    clear_env(monkeypatch)
    prompts_dir = tmp_path / "prompts"
    prompts_dir.mkdir()
    create_prompt_assets(prompts_dir)
    ids_file = tmp_path / "ids.txt"
    ids_file.write_text("201\n202\n203\n", encoding="utf-8")
    xlsx_path = tmp_path / "batch.xlsx"

    monkeypatch.setenv("PRIMARY_MODEL", "primary")
    monkeypatch.setenv("AUDIT_MODEL", "audit")
    monkeypatch.setenv("RANKING_MODEL", "ranking")
    monkeypatch.setenv("MAIN_DATABASE_URL", "postgresql://main")
    monkeypatch.setenv("RUNTIME_DATABASE_URL", "postgresql://runtime")
    monkeypatch.setenv("PROMPTS_DIR", str(prompts_dir))
    monkeypatch.setenv("OUTPUT_DIR", str(tmp_path / "output"))
    monkeypatch.setenv("EVENT_MAX_CONCURRENCY", "2")

    active_counter = {"value": 0}
    max_seen_counter = {"value": 0}
    lock = threading.Lock()
    monkeypatch.setattr(
        "buratino.cli.main.build_app",
        lambda settings: ControlledFakeApp(
            active_counter=active_counter,
            max_seen_counter=max_seen_counter,
            lock=lock,
            failures={202},
        ),
    )

    exit_code = run(["verify-list", str(ids_file), "--xlsx", str(xlsx_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "Processed ids: 3" in captured.out
    assert "Successful: 2" in captured.out
    assert "Failed: 1" in captured.out
    assert (tmp_path / "output" / "event_201.json").exists()
    assert not (tmp_path / "output" / "event_202.json").exists()
    assert (tmp_path / "output" / "event_203.json").exists()
